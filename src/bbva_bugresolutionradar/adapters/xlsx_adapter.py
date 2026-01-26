from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from bbva_bugresolutionradar.adapters.filesystem import FilesystemAdapter
from bbva_bugresolutionradar.adapters.utils import to_date, to_str
from bbva_bugresolutionradar.domain.enums import Severity, Status
from bbva_bugresolutionradar.domain.models import ObservedIncident


class XlsxAdapter(FilesystemAdapter):
    """
    XLSX adapter de filesystem.
    Patrón idéntico a FilesystemCSVAdapter / FilesystemJSONAdapter:
      - recibe (source_id, assets_root)
      - assets_root apunta a ASSETS_DIR
    """

    def __init__(self, source_id: str, assets_root: str) -> None:
        super().__init__(source_id, assets_root)

    def read(self) -> list[ObservedIncident]:
        out: list[ObservedIncident] = []

        # ✅ Sin env_optional(): valores por defecto seguros
        ignore: set[str] = set()
        preferred_sheet = "Reportes"

        for path in sorted(self.assets_dir().glob("*.xlsx")):
            if path.name in ignore:
                continue

            # Caso específico: Canales Digitales Enterprise.xlsx
            if "Canales Digitales Enterprise" in path.name:
                out.extend(self._read_canales_enterprise(path, preferred_sheet))

        return out

    def _read_canales_enterprise(self, path: Path, preferred_sheet: str) -> list[ObservedIncident]:
        wb = openpyxl.load_workbook(path, data_only=True)

        ws = wb[preferred_sheet] if preferred_sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [to_str(h) for h in rows[0]]

        def idx(contains: str) -> Optional[int]:
            needle = contains.lower()
            for i, h in enumerate(headers):
                if h and needle in h.lower():
                    return i
            return None

        col_id = idx("id de reporte")
        col_opened = idx("fecha de incidente")
        col_status = idx("estatus")
        col_sev = idx("criticidad")
        col_title = idx("descripción")
        col_feature = idx("funcionalidad")
        col_product = idx("tema") or idx("canal")

        if col_id is None or col_opened is None:
            return []

        observed_at = datetime.now().astimezone()
        out: list[ObservedIncident] = []

        for r in rows[1:]:
            source_key = to_str(r[col_id])
            opened_at = to_date(r[col_opened])

            if not source_key or opened_at is None:
                continue

            status_raw = to_str(r[col_status]) if col_status is not None else None
            sev_raw = to_str(r[col_sev]) if col_sev is not None else None

            status = _map_status(status_raw)
            severity = _map_severity(sev_raw)

            title = to_str(r[col_title]) if col_title is not None else None
            feature = to_str(r[col_feature]) if col_feature is not None else None
            product = to_str(r[col_product]) if col_product is not None else None

            out.append(
                ObservedIncident(
                    source_id=self.source_id(),
                    source_key=source_key,
                    observed_at=observed_at,
                    title=title or "",
                    status=status,
                    severity=severity,
                    opened_at=opened_at,
                    closed_at=None,
                    updated_at=None,
                    clients_affected=None,
                    product=product,
                    feature=feature,
                    resolution_type=None,
                )
            )

        return out


def _map_status(raw: str | None) -> Status:
    if not raw:
        return Status.UNKNOWN
    s = raw.strip().lower()
    if "abiert" in s or "open" in s:
        return Status.OPEN
    if "cerr" in s or "close" in s:
        return Status.CLOSED
    if "progreso" in s or "progress" in s:
        return Status.IN_PROGRESS
    if "bloque" in s or "block" in s:
        return Status.BLOCKED
    return Status.UNKNOWN


def _map_severity(raw: str | None) -> Severity:
    if not raw:
        return Severity.UNKNOWN
    s = raw.strip().lower()
    if "crit" in s:
        return Severity.CRITICAL
    if "alta" in s or "high" in s:
        return Severity.HIGH
    if "media" in s or "med" in s:
        return Severity.MEDIUM
    if "baja" in s or "low" in s:
        return Severity.LOW
    return Severity.UNKNOWN